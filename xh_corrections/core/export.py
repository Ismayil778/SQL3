"""
Export to Excel (openpyxl) — 3 sheets:

  Sheet 1 "Müxabirləşmələr" — Types 3-4 corrective entries
  Sheet 2 "Xitam"            — Closed policies with empty AMOUNT for manual fill
  Sheet 3 "Xülasə"           — Summary statistics
"""
from datetime import datetime
from pathlib import Path

import openpyxl  # noqa: runtime dependency
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: runtime dependency
from openpyxl.utils import get_column_letter  # noqa: runtime dependency


XH_RED_HEX    = "C8102E"
XH_WHITE_HEX  = "FFFFFF"
XH_LIGHT_HEX  = "F5F5F5"
XH_BORDER_HEX = "E0E0E0"
XH_AMBER_HEX  = "FFF3CD"   # xitam sheet background
XH_AMBER_BDR  = "E8A020"   # xitam border


def _side(color: str = "CCCCCC") -> Side:
    return Side(style="thin", color=color)


def _thin_border(color: str = "CCCCCC") -> Border:
    s = _side(color)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=XH_RED_HEX)


def _amber_fill() -> PatternFill:
    return PatternFill("solid", fgColor=XH_AMBER_HEX)


def export_to_excel(
    corrections: list[dict],
    xitam_list: list[dict],
    total_policies: int,
    report_date_str: str,   # 'YYYYMMDD'
    output_path: str = "",
) -> str:
    """
    Write Excel file to output_path and return the absolute path.

    corrections: list of DT/KT/AMOUNT/Policy_Number/Months rows
    xitam_list:  list of {policy_number, policy_complete_date, policy_id}
    """
    rd = report_date_str
    report_date_fmt = f"{rd[6:8]}.{rd[4:6]}.{rd[:4]}"

    if not output_path:
        output_path = str(Path.cwd() / f"korreksiyalar_{report_date_str}.xlsx")

    wb = openpyxl.Workbook()

    _write_entries_sheet(wb, corrections)
    _write_xitam_sheet(wb, xitam_list, report_date_fmt)
    _write_summary_sheet(wb, corrections, xitam_list, total_policies, report_date_fmt)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    return str(Path(output_path).resolve())


# ---------------------------------------------------------------------------
# Sheet 1 — Müxabirləşmələr
# ---------------------------------------------------------------------------

def _write_entries_sheet(
    wb: openpyxl.Workbook,
    corrections: list[dict],
) -> None:
    ws = wb.create_sheet("Müxabirləşmələr")

    headers    = ["DT", "KT", "AMOUNT", "Siyasət / Полис", "Ay / Месяцев"]
    col_widths = [14,   14,   16,       24,                 14]

    hdr_font  = Font(name="Calibri", bold=True, color=XH_WHITE_HEX, size=11)
    hdr_fill  = _header_fill()
    hdr_align = Alignment(horizontal="center", vertical="center")

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = _thin_border("AAAAAA")
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    sorted_rows = sorted(corrections, key=lambda r: r.get("Policy_Number", ""))

    light_fill  = PatternFill("solid", fgColor=XH_LIGHT_HEX)
    white_fill  = PatternFill("solid", fgColor=XH_WHITE_HEX)
    data_font   = Font(name="Calibri", size=10)
    center_aln  = Alignment(horizontal="center", vertical="center")
    left_aln    = Alignment(horizontal="left",   vertical="center")
    right_aln   = Alignment(horizontal="right",  vertical="center")

    for ri, row in enumerate(sorted_rows, start=2):
        fill   = white_fill if ri % 2 == 0 else light_fill
        values = [
            row.get("DT", ""),
            row.get("KT", ""),
            row.get("AMOUNT"),
            row.get("Policy_Number", ""),
            row.get("Months", ""),
        ]
        aligns = [center_aln, center_aln, right_aln, left_aln, center_aln]

        for ci, (val, aln) in enumerate(zip(values, aligns), start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = data_font
            cell.fill      = fill
            cell.alignment = aln
            cell.border    = _thin_border(XH_BORDER_HEX)
            if ci == 3:
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Sheet 2 — Xitam (closed policies)
# ---------------------------------------------------------------------------

def _write_xitam_sheet(
    wb: openpyxl.Workbook,
    xitam_list: list[dict],
    report_date_fmt: str,
) -> None:
    ws = wb.create_sheet("Xitam")

    # Note row
    note_font = Font(name="Calibri", bold=True, size=10, color="7B5A00")
    note_fill = PatternFill("solid", fgColor="FFF3CD")
    note_cell = ws.cell(
        row=1, column=1,
        value=(
            "QEYD / ПРИМЕЧАНИЕ: AMOUNT sütunu əl ilə doldurulmalıdır — "
            "məbləği XalqLife sistemindən götürün / "
            "Столбец AMOUNT заполняется вручную из системы XalqLife"
        ),
    )
    note_cell.font      = note_font
    note_cell.fill      = note_fill
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 36

    headers    = ["Siyasət / Полис", "Bağlanma tarixi / Дата закрытия", "AMOUNT"]
    col_widths = [26,                 28,                                  18]

    hdr_font  = Font(name="Calibri", bold=True, color=XH_WHITE_HEX, size=11)
    hdr_fill  = PatternFill("solid", fgColor="8B6914")   # dark amber header
    hdr_align = Alignment(horizontal="center", vertical="center")

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = _thin_border("8B6914")
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[2].height = 22

    amber_fill  = _amber_fill()
    data_font   = Font(name="Calibri", size=10)
    center_aln  = Alignment(horizontal="center", vertical="center")
    left_aln    = Alignment(horizontal="left",   vertical="center")

    for ri, row in enumerate(xitam_list, start=3):
        complete_date = row.get("policy_complete_date")
        date_str = ""
        if complete_date is not None:
            try:
                date_str = complete_date.strftime("%d.%m.%Y")
            except AttributeError:
                date_str = str(complete_date)

        values = [
            row.get("policy_number", ""),
            date_str,
            None,   # AMOUNT — intentionally empty for manual fill
        ]
        aligns = [left_aln, center_aln, center_aln]

        for ci, (val, aln) in enumerate(zip(values, aligns), start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = data_font
            cell.fill      = amber_fill
            cell.alignment = aln
            cell.border    = _thin_border(XH_AMBER_BDR)
            if ci == 3:
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Sheet 3 — Xülasə (summary)
# ---------------------------------------------------------------------------

def _write_summary_sheet(
    wb: openpyxl.Workbook,
    corrections: list[dict],
    xitam_list: list[dict],
    total_policies: int,
    report_date_fmt: str,
) -> None:
    ws = wb.create_sheet("Xülasə")

    policy_set   = {r["Policy_Number"] for r in corrections}
    total_amount = sum(r.get("AMOUNT", 0) for r in corrections if r.get("DT") == "84.1.1.")
    now_fmt      = datetime.now().strftime("%d.%m.%Y %H:%M")

    rows = [
        ("Hesab tarixi / Дата расчёта",                              report_date_fmt),
        ("Cəmi aktiv polislər / Активных полисов (XalqLife)",        total_policies),
        ("Korreksiya olan polislər / Полисов с корректировками",     len(policy_set)),
        ("Bağlı polislər (xitam) / Закрытых полисов",               len(xitam_list)),
        ("Cəmi müxabirləşmələr / Всего проводок",                   len(corrections)),
        ("Cəmi məbləğ (84.1.1.) / Общая сумма (84.1.1.)",           total_amount),
        ("Yaradılma vaxtı / Дата создания",                         now_fmt),
    ]

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 24

    title_font  = Font(name="Calibri", bold=True, color=XH_WHITE_HEX, size=12)
    title_fill  = _header_fill()
    label_font  = Font(name="Calibri", size=10, color="2D2D2D")
    value_font  = Font(name="Calibri", bold=True, size=10)
    light_fill  = PatternFill("solid", fgColor=XH_LIGHT_HEX)
    white_fill  = PatternFill("solid", fgColor=XH_WHITE_HEX)

    title_cell = ws.cell(
        row=1, column=1,
        value="Xalq Həyat ASC — Korreksiya Xülasəsi / Сводка корректировок",
    )
    ws.merge_cells("A1:B1")
    title_cell.font      = title_font
    title_cell.fill      = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for i, (label, value) in enumerate(rows, start=2):
        fill = light_fill if i % 2 == 0 else white_fill

        lc = ws.cell(row=i, column=1, value=label)
        lc.font      = label_font
        lc.fill      = fill
        lc.alignment = Alignment(horizontal="left",  vertical="center")
        lc.border    = _thin_border(XH_BORDER_HEX)

        vc = ws.cell(row=i, column=2, value=value)
        vc.font      = value_font
        vc.fill      = fill
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.border    = _thin_border(XH_BORDER_HEX)
        if isinstance(value, float):
            vc.number_format = "#,##0.00"

        ws.row_dimensions[i].height = 20

    ws.sheet_view.showGridLines = False
