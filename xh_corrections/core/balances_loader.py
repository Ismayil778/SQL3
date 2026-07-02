"""
Loads account balances from Excel file as of period start (e.g., 31.03.2026).

Expected columns (auto-detected by keyword):
  - Policy number: "Polis Nömrəsi", "Policy", "Номер"
  - Account:       "Hesab", "Счет", "Account"
  - Balance:       "Qalıq Məbləğ", "Məbləğ", "Остаток", "Сумма", "Balance"

Account codes are normalised: "78. 1. 1.1." → "78.1.1.1"

Returns:
    balances: dict[policy_number, dict[account, float]]
    total_policies: int
"""
import re
from openpyxl import load_workbook


def _normalize_account(raw) -> str:
    """Remove all whitespace from account string. Strip trailing dots."""
    s = re.sub(r'\s+', '', str(raw).strip())
    return s.rstrip('.')


def _find_col(headers: list[str], keywords: list[str]) -> int | None:
    for i, h in enumerate(headers):
        h_lower = h.lower()
        if any(kw in h_lower for kw in keywords):
            return i
    return None


def load_balances(file_path: str) -> tuple[dict[str, dict[str, float]], int]:
    """
    Load account balances from Excel file.

    Returns:
        (balances, total_policies)
        balances: {policy_number: {account: float}}
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise ValueError("Excel fayl boşdur / Excel файл пустой")

    # Find first non-empty row as header
    header_idx = 0
    header_row: list[str] = []
    for i, row in enumerate(all_rows):
        non_empty = [str(c).strip() for c in row if c is not None]
        if non_empty:
            header_row = [str(c).strip() if c is not None else "" for c in row]
            header_idx = i
            break

    if not header_row:
        raise ValueError("Başlıq sətri tapılmadı / Не найдена строка заголовков")

    policy_col = _find_col(header_row, ["polis", "policy", "nömrə", "номер", "полис"])
    account_col = _find_col(header_row, ["hesab", "счет", "account", "schet"])
    amount_col = _find_col(
        header_row,
        ["məbləğ", "qalıq", "остаток", "сумма", "balance", "amount", "balans"],
    )

    if policy_col is None:
        raise ValueError(
            f"Polis sütunu tapılmadı. Başlıqlar: {header_row}\n"
            "Gözlənilən: 'Polis Nömrəsi', 'Policy Number', 'Номер полиса'"
        )
    if account_col is None:
        raise ValueError(
            f"Hesab sütunu tapılmadı. Başlıqlar: {header_row}\n"
            "Gözlənilən: 'Hesab', 'Account', 'Счет'"
        )
    if amount_col is None:
        raise ValueError(
            f"Məbləğ sütunu tapılmadı. Başlıqlar: {header_row}\n"
            "Gözlənilən: 'Qalıq Məbləğ', 'Остаток', 'Balance'"
        )

    balances: dict[str, dict[str, float]] = {}

    for row in all_rows[header_idx + 1:]:
        if len(row) <= max(policy_col, account_col, amount_col):
            continue

        policy_raw = row[policy_col]
        account_raw = row[account_col]
        amount_raw = row[amount_col]

        if policy_raw is None or account_raw is None:
            continue

        policy = str(policy_raw).strip()
        account = _normalize_account(account_raw)

        if not policy or not account:
            continue

        # Skip rows that look like headers repeated inside data
        if policy.lower() in ("polis", "policy", "nömrəsi", "номер"):
            continue

        try:
            amount = float(str(amount_raw).replace(',', '.')) if amount_raw is not None else 0.0
        except (TypeError, ValueError):
            amount = 0.0

        if policy not in balances:
            balances[policy] = {}
        balances[policy][account] = balances[policy].get(account, 0.0) + amount

    return balances, len(balances)
